import datetime

from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse, HttpRequest

from polls.models import Subject, Teacher, User
from polls.utils import Captcha, gen_random_code, gen_md5_digest


# Create your views here.

def show_subjects(request):
    subjects = Subject.objects.all().order_by('no')
    return render(request, 'subjects.html', {'subjects': subjects})


def show_teachers(request, sno):
    try:
        # sno = int(request.GET.get('sno'))
        teachers = []
        if sno:
            subject = Subject.objects.only('name').get(no=sno)
            teachers = Teacher.objects.filter(subject=subject).order_by('no')
        return render(request, 'teachers.html', {
            'subject': subject,
            'teachers': teachers
        })
    except (ValueError, Subject.DoesNotExist):
        return redirect('/')


# def praise_or_criticize(request, tno):
#     """好评/差评"""
#     try:
#         teacher = Teacher.objects.get(no=tno)
#         if request.path.startswith('/praise'):
#             teacher.good_count += 1
#             count = teacher.good_count
#         else:
#             teacher.bad_count += 1
#             count = teacher.bad_count
#         teacher.save()
#         data = {'code': '20000', 'mesg': '操作成功', 'count': count}
#     except Teacher.DoesNotExist:
#         data = {'code': 20001, 'mesg': '操作失败'}
#     return JsonResponse(data)


def praise_or_criticize(request: HttpRequest, tno: int) -> HttpResponse:
    if request.session.get('userid'):
        try:
            teacher = Teacher.objects.get(no=tno)
            if request.path.startswith('/praise/'):
                teacher.good_count += 1
                count = teacher.good_count
            else:
                teacher.bad_count += 1
                count = teacher.bad_count
            teacher.save()
            data = {'code': 20000, 'mesg': '投票成功', 'count': count}
        except (ValueError, Teacher.DoesNotExist):
            data = {'code': 20001, 'mesg': '投票失败'}
    else:
        data = {'code': 20002, 'mesg': '请先登录'}
    return JsonResponse(data)


def login(request: HttpRequest) -> HttpResponse:
    hint = ''
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        captcha = request.POST.get('captcha')
        # 将用户提供的验证码与session中存储的验证码进行比对
        if captcha.upper() == request.session['captcha'].upper():
            if username and password:
                password = gen_md5_digest(password)
                user = User.objects.filter(username=username, password=password).first()
                if user:
                    # 在session中存入userid,username
                    request.session['userid'] = user.no
                    request.session['username'] = user.username
                    last_visit = datetime.datetime.now()
                    # 将登录信息存储cookie
                    response = redirect('/')
                    response.set_cookie('last_visit', last_visit)
                    return response
                else:
                    hint = '用户名或密码错误'
            else:
                hint = '请输入有效的用户名和密码'
        else:
            hint = '验证码错误'
    return render(request, 'login.html', {'hint': hint})


def get_captcha(request: HttpRequest) -> HttpResponse:
    # 生成验证码字符串
    captcha_text = gen_random_code()
    # 将验证码字符串存入session
    request.session['captcha'] = captcha_text
    image_data = Captcha.instance().generate(captcha_text)
    return HttpResponse(image_data, content_type='image/png')


def logout(request):
    """注销"""
    # 清除session中的数据
    request.session.flush()
    rep = redirect('/')
    # 清除cookie信息
    # rep.delete_cookie('last_visit')
    return rep


def export_teachers_excel(request):
    from openpyxl import Workbook
    from io import BytesIO
    from django.utils.http import urlquote

    # 查询所有老师的信息
    queryset = Teacher.objects.all()

    # 生成一个工作簿（即一个Excel文件）
    wb = Workbook()
    wb.encoding = 'utf-8'
    sheet1 = wb.active  # 获取第一个工作表（sheet1）
    sheet1.title = '老师信息表'

    colnames = ('姓名', '介绍', '好评数', '差评数', '学科')
    for index, name in enumerate(colnames):
        sheet1.cell(row=1, column=index + 1).value = colnames[index]

    # 逐行将老师信息写入工作表
    for teacher in queryset:
        teacher_info = [teacher.name, teacher.intro, teacher.good_count, teacher.bad_count, teacher.subject.name]
        current_row = sheet1.max_row + 1  # 当前行号
        for i in range(1, len(teacher_info) + 1):
            sheet1.cell(row=current_row, column=i).value = teacher_info[i - 1]
    # 准备写入IO中
    output = BytesIO()
    wb.save(output)  # 将excel文件内容保存到IO中
    output.seek(0)  # 重新定位到开始
    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    create_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    file_name = f'老师信息表{create_time}.xls'
    file_name = urlquote(file_name)  # 使用urlquote()方法解决中文无法使用问题
    response['Content-Disposition'] = f'attachment; filename={file_name}'
    # response.write(output.getvalue())	 # 在设置HttpResponse的类型时，如果给了值，可以不写这句
    return response


def get_teachers_data(request):
    queryset = Teacher.objects.all()
    names = [teacher.name for teacher in queryset]
    good_counts = [teacher.good_count for teacher in queryset]
    bad_counts = [teacher.bad_count for teacher in queryset]
    return JsonResponse({'names': names, 'good': good_counts, 'bad': bad_counts})


def get_teachers_chart(request):
    return render(request, 'teacher_chart.html')
